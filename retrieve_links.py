#!/usr/bin/env python3

print("Running script...")

"""
WorldCat Search API v2 - ISBN naar OCN checker (ebooks)

Gebaseerd op de officiele OpenAPI-spec:
https://developer.api.oclc.org/wcv2

Workflow per ISBN:
    /brief-bibs?q=bn:"{isbn}"&heldBySymbol=...
    -> OCNs waar jouw instelling een holding op heeft
    -> 1 OCN: hyperlink in kolom "Link"
    -> Meerdere OCNs: kommagescheiden in kolom "OCN", hyperlinks elk in eigen kolom

Vereisten:
    pip install requests pandas openpyxl tqdm python-dotenv

Stel eerst je API-sleutels in via:
    python config.py

Benodigde scope voor je WSKey (developer.api.oclc.org):
    wcapi:view_institution_holdings

Gebruik:
    Leg input.xlsx in dezelfde map als dit script en voer uit:
    python retrieve_links.py
"""

import sys
import time
import base64
from pathlib import Path
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from tqdm import tqdm
import config as cfg

# ---------------------------------------------------------------------------
# Instellingen
# ---------------------------------------------------------------------------
TOKEN_URL    = "https://oauth.oclc.org/token"
API_BASE     = "https://americas.discovery.api.oclc.org/worldcat/search/v2"
BRIEF_BIB    = f"{API_BASE}/brief-bibs"
BIB_HOLDINGS = f"{API_BASE}/bibs-holdings"

DELAY_SECONDS = 0.3

SCRIPT_DIR = Path(__file__).parent
INPUT_FILE = SCRIPT_DIR / "input.xlsx"

DEFAULT_ISBN_COL   = "ISBN"
DEFAULT_OCN_COL    = "OCN"
DEFAULT_STATUS_COL = "Status"

WORLDCAT_BASE_URL  = "https://eur.on.worldcat.org/oclc/"

# ---------------------------------------------------------------------------
# Authenticatie
# ---------------------------------------------------------------------------

class TokenManager:
    """Beheert het OAuth-token en vernieuwt automatisch bij verlopen."""

    def __init__(self, wskey: str, secret: str):
        self.wskey  = wskey
        self.secret = secret
        self._token = None

    def get(self) -> str:
        if not self._token:
            self._token = self._fetch()
        return self._token

    def refresh(self) -> str:
        self._token = self._fetch()
        return self._token

    def _fetch(self) -> str:
        credentials = base64.b64encode(f"{self.wskey}:{self.secret}".encode()).decode()
        resp = requests.post(
            TOKEN_URL,
            headers={
                "Authorization": f"Basic {credentials}",
                "Content-Type": "application/x-www-form-urlencoded",
            },
            data={"grant_type": "client_credentials", "scope": "wcapi"},
            timeout=30,
        )
        if resp.status_code != 200:
            print(f"\n[FOUT] Token ophalen mislukt ({resp.status_code}): {resp.text}")
            sys.exit(1)
        return resp.json()["access_token"]


# ---------------------------------------------------------------------------
# API-hulpfunctie
# ---------------------------------------------------------------------------

def _get(url: str, params: dict, token_mgr: TokenManager) -> dict:
    """GET met automatische token-refresh bij 401."""
    for attempt in range(2):
        try:
            resp = requests.get(
                url,
                params=params,
                headers={
                    "Authorization": f"Bearer {token_mgr.get()}",
                    "Accept": "application/json",
                },
                timeout=30,
            )
        except requests.RequestException as exc:
            return {"status": 0, "body": {}, "error": str(exc)}

        if resp.status_code == 401 and attempt == 0:
            token_mgr.refresh()
            continue

        try:
            body = resp.json()
        except Exception:
            body = {}
        return {"status": resp.status_code, "body": body}

    return {"status": 401, "body": {}}


# ---------------------------------------------------------------------------
# Verwerking per ISBN
# ---------------------------------------------------------------------------

def process_isbn(isbn: str, symbol: str, token_mgr: TokenManager) -> dict:
    """
    Zoek via brief-bibs alle OCNs op waar jouw instelling een holding op heeft.
    Geeft {"ocns": [...], "status": "..."} terug.
    """
    result = _get(
        BRIEF_BIB,
        {
            "q":            f'bn:"{isbn}"',
            "heldBySymbol": symbol,
            "itemSubType":  "book-digital",
            "limit":        50,
        },
        token_mgr,
    )

    if "error" in result:
        return {"ocn": "", "status": f"API-fout: {result['error']}"}
    if result["status"] == 400:
        msg = result["body"].get("message", result["body"])
        return {"ocn": "", "status": f"API-fout: Ongeldig verzoek (400): {msg}"}
    if result["status"] != 200:
        return {"ocn": "", "status": f"API-fout: HTTP {result['status']}"}

    records = result["body"].get("briefRecords", [])
    ocns = [str(r["oclcNumber"]) for r in records if r.get("oclcNumber")]
 
    if not ocns:
        return {"ocns": [], "status": "Geen OCN met holding gevonden"}
 
    if len(ocns) == 1:
        status = "Holding gevonden (1 OCN)"
    else:
        status = f"Holding gevonden ({len(ocns)} OCNs)"
 
    return {"ocns": ocns, "status": status}
 
 # ---------------------------------------------------------------------------
# Opmaak: hyperlink stijl (blauw + onderstreept) toepassen via openpyxl
# ---------------------------------------------------------------------------
 
def apply_hyperlink_style(ws, col_letter: str, row: int):
    cell = ws[f"{col_letter}{row}"]
    cell.font = Font(color="0563C1", underline="single")
 
# ---------------------------------------------------------------------------
# Hoofdverwerking
# ---------------------------------------------------------------------------

def main():
    print("=" * 54)
    print("  WorldCat Search API v2 - ISBN ebook holding checker")
    print("=" * 54)

    if not INPUT_FILE.exists():
        print(f"\n[FOUT] Bestand niet gevonden: {INPUT_FILE}")
        print("Zet input.xlsx in dezelfde map als dit script.")
        sys.exit(1)

    if not cfg.credentials_exist():
        print("\n[FOUT] Bestand .env niet gevonden.")
        print("Maak een .env-bestand aan in dezelfde map als het script met WSKEY, WSKEY_SECRET en INSTITUTION_SYMBOL.")
        sys.exit(1)

    try:
        creds = cfg.load()
    except ValueError as exc:
        print(f"\n[FOUT] {exc}")
        sys.exit(1)

    token_mgr = TokenManager(creds["WSKEY"], creds["WSKEY_SECRET"])
    symbol    = creds["INSTITUTION_SYMBOL"]

    print(f"\ninput.xlsx gevonden!")
    df = pd.read_excel(INPUT_FILE, dtype=str)

    if DEFAULT_ISBN_COL not in df.columns:
        print(f"[FOUT] Kolom '{DEFAULT_ISBN_COL}' niet gevonden.")
        print(f"Beschikbare kolommen: {list(df.columns)}")
        sys.exit(1)

    df[DEFAULT_OCN_COL]    = ""
    df[DEFAULT_STATUS_COL] = ""
    df["Link"]             = ""

    total     = len(df)
    found     = 0
    multi     = 0
    not_found = 0
    errors    = 0
    
    # Bijhouden welke rijen meerdere OCNs hebben voor extra kolommen later
    multi_ocn_rows = {}   # idx -> [ocn1, ocn2, ...]
    max_ocns = 1

    print(f"{total} rijen verwerken...\n")

    for idx, row in tqdm(df.iterrows(), total=total, unit="ISBN"):
        raw_isbn = str(row[DEFAULT_ISBN_COL]).strip()
 
        if not raw_isbn or raw_isbn.lower() in ("nan", "none", ""):
            df.at[idx, DEFAULT_STATUS_COL] = "Geen ISBN"
            continue
 
        isbn   = raw_isbn.replace("-", "").replace(" ", "")
        result = process_isbn(isbn, symbol, token_mgr)
        time.sleep(DELAY_SECONDS)
 
        ocns   = result["ocns"]
        status = result["status"]
 
        df.at[idx, DEFAULT_STATUS_COL] = status
 
        if not ocns:
            continue
 
        # OCN-kolom: altijd alle OCNs kommagescheiden
        df.at[idx, DEFAULT_OCN_COL] = ", ".join(ocns)
 
        if len(ocns) == 1:
            # Eén OCN: hyperlink in kolom "Link"
            df.at[idx, "Link"] = f'=HYPERLINK("{WORLDCAT_BASE_URL}{ocns[0]}","{ocns[0]}")'
        else:
            # Meerdere OCNs: hyperlinks komen elk in eigen kolom (Link 1, Link 2, ...)
            multi_ocn_rows[idx] = ocns
            max_ocns = max(max_ocns, len(ocns))
 
        s = status
        if s.startswith("Holding gevonden"):
            found += 1
            if "OCNs)" in s:
                multi += 1
        elif s == "Geen OCN met holding gevonden":
            not_found += 1
        elif s.startswith("API-fout"):
            errors += 1
 
    # Voeg Link-kolommen toe voor meerdere OCNs
    if multi_ocn_rows:
        for i in range(1, max_ocns + 1):
            col = f"Link {i}"
            if col not in df.columns:
                df[col] = ""
        for idx, ocns in multi_ocn_rows.items():
            df.at[idx, "Link"] = ""   # leeglaten, per-kolom links worden gebruikt
            for i, ocn in enumerate(ocns, 1):
                df.at[idx, f"Link {i}"] = f'=HYPERLINK("{WORLDCAT_BASE_URL}{ocn}","{ocn}")'
 
    # Opslaan
    output_path = SCRIPT_DIR / "output.xlsx"
    df.to_excel(output_path, index=False)
 
    # Hyperlink opmaak toepassen via openpyxl
    wb = load_workbook(output_path)
    ws = wb.active
 
    # Zoek kolomletters op basis van koptekst
    header = {cell.value: cell.column_letter for cell in ws[1]}
 
    for excel_row in range(2, len(df) + 2):
        # Kolom "Link" (enkelvoudig)
        if "Link" in header:
            cell = ws[f"{header['Link']}{excel_row}"]
            if cell.value and str(cell.value).startswith("=HYPERLINK"):
                cell.font = Font(color="0563C1", underline="single")
 
        # Kolommen "Link 1", "Link 2", ...
        for i in range(1, max_ocns + 1):
            col_name = f"Link {i}"
            if col_name in header:
                cell = ws[f"{header[col_name]}{excel_row}"]
                if cell.value and str(cell.value).startswith("=HYPERLINK"):
                    cell.font = Font(color="0563C1", underline="single")
 
    wb.save(output_path)
 
    print(f"""
Totaal verwerkt                 : {total}
Holding gevonden                : {found}  (waarvan {multi} met meerdere OCNs)
Geen OCN met holding gevonden   : {not_found}
Fouten                          : {errors}
 
Resultaat opgeslagen in         : {output_path}
""")
 
 
if __name__ == "__main__":
    main()
 
